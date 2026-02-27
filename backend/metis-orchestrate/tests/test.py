from superuser_login import get_superuser_token
from tenant_onboarding import TenantOnboarding
from tenant_email_verification import TenantEmailVerification
from request_otp import RequestOTP
from validate_login_otp import ValidateOTP
from get_tenant_roles import TenantRoles
from tenant_users import TenantUsers

import json
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

class TestAutomation:
    superuser_token = None
    tenant_count = 2
    tenant_slots = {}
    tenants = {}
    owners = {}
    otp = "999999"
    test_cases_sheets = {}

    def __init__(self):
        self.tenant_slots = {
            idx: {"tenant": {}, "owner_user": {}, "owners": {}, "users": {}, "roles": []}
            for idx in range(1, self.tenant_count + 1)
        }
        self.tenants = {}
        self.owners = {}
        self.test_cases_sheets = {}
        self.superuser_token = get_superuser_token()
        self.start_test()

    def _beautify_json_cell(self, value):
        if value is None:
            return ""

        if isinstance(value, (dict, list)):
            return json.dumps(value, indent=2, ensure_ascii=False)

        if isinstance(value, str):
            raw = value.strip()
            if (raw.startswith("{") and raw.endswith("}")) or (raw.startswith("[") and raw.endswith("]")):
                try:
                    parsed = json.loads(raw)
                    return json.dumps(parsed, indent=2, ensure_ascii=False)
                except Exception:
                    return value

        return value


    def _export_excel(self, filename, sheets):
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        row_fill_odd = PatternFill(fill_type="solid", fgColor="F7FBFF")
        row_fill_even = PatternFill(fill_type="solid", fgColor="EAF2FB")
        pass_fill = PatternFill(fill_type="solid", fgColor="D9F2D9")
        fail_fill = PatternFill(fill_type="solid", fgColor="FADBD8")
        thin_side = Side(style="thin", color="D9D9D9")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            summary_rows = []
            for case_name, data in sheets.items():
                total_cases = len(data or [])
                passed_count = 0
                failed_count = 0
                for case in (data or []):
                    if not isinstance(case, dict):
                        continue
                    result_value = str(case.get("result", "")).strip().lower()
                    if result_value == "passed":
                        passed_count += 1
                    elif result_value == "failed":
                        failed_count += 1
                summary_rows.append(
                    {
                        "Case Name": case_name,
                        "No of Cases": total_cases,
                        "No of Passed": passed_count,
                        "No of Failed": failed_count,
                    }
                )

            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            summary_ws = writer.book["Summary"]
            summary_ws.freeze_panes = "A2"
            summary_max_row = summary_ws.max_row
            summary_max_col = summary_ws.max_column

            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx in range(2, summary_max_row + 1):
                default_fill = row_fill_odd if row_idx % 2 else row_fill_even
                for col_idx in range(1, summary_max_col + 1):
                    cell = summary_ws.cell(row=row_idx, column=col_idx)
                    cell.fill = default_fill
                    cell.border = cell_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for col_idx in range(1, summary_max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, summary_max_row + 1):
                    cell_value = summary_ws.cell(row=row_idx, column=col_idx).value
                    max_len = max(max_len, len(str(cell_value or "")))
                summary_ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

            for sheet_name, data in sheets.items():
                safe_sheet_name = (sheet_name or "Sheet1")[:31]
                formatted_data = []
                for row in (data or []):
                    if isinstance(row, dict):
                        formatted_row = dict(row)
                        if "payload" in formatted_row:
                            formatted_row["payload"] = self._beautify_json_cell(formatted_row.get("payload"))
                        if "response" in formatted_row:
                            formatted_row["response"] = self._beautify_json_cell(formatted_row.get("response"))
                        formatted_data.append(formatted_row)
                    else:
                        formatted_data.append(row)

                df = pd.DataFrame(formatted_data)
                df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                worksheet = writer.book[safe_sheet_name]
                worksheet.freeze_panes = "A2"

                max_row = worksheet.max_row
                max_col = worksheet.max_column

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = cell_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                header_to_col = {}
                for col_idx in range(1, max_col + 1):
                    header_value = worksheet.cell(row=1, column=col_idx).value
                    header_to_col[str(header_value).strip().lower()] = col_idx
                result_col = header_to_col.get("result")
                payload_col = header_to_col.get("payload")
                response_col = header_to_col.get("response")
                json_cols = {col for col in [payload_col, response_col] if col}

                for row_idx in range(2, max_row + 1):
                    default_fill = row_fill_odd if row_idx % 2 else row_fill_even
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = default_fill
                        cell.border = cell_border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                    has_json_content = False
                    for json_col in json_cols:
                        json_cell = worksheet.cell(row=row_idx, column=json_col)
                        json_text = str(json_cell.value or "").strip()
                        if json_text:
                            has_json_content = True
                            json_cell.font = Font(name="Consolas", size=10)
                            json_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                    if has_json_content:
                        worksheet.row_dimensions[row_idx].height = 90

                    if result_col:
                        result_cell = worksheet.cell(row=row_idx, column=result_col)
                        result_value = str(result_cell.value).strip().lower()
                        if result_value == "passed":
                            result_cell.fill = pass_fill
                        elif result_value == "failed":
                            result_cell.fill = fail_fill
                        result_cell.font = Font(bold=True)
                        result_cell.alignment = Alignment(horizontal="center", vertical="center")

                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for row_idx in range(1, max_row + 1):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        max_len = max(max_len, len(str(cell_value or "")))
                    if col_idx in json_cols:
                        worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 30), 90)
                    else:
                        worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    def tenant__owner_user_email(self, tenant_index):
        return self.tenant_slots.get(tenant_index, {}).get("owner_user", {}).get("email")

    def _owner_user_password(self, tenant_index):
        return self.tenant_slots.get(tenant_index, {}).get("owner_user", {}).get("password")

    def _tenant_id(self, tenant_index):
        return self.tenant_slots.get(tenant_index, {}).get("tenant", {}).get("id")

    def _logged_in_tenant_owner_user_access_token(self, tenant_index):
        owners = self.tenant_slots.get(tenant_index, {}).get("owners", {})
        if owners:
            latest_user_data = next(reversed(owners.values()))
            return latest_user_data.get("access")
        return None

    def _store_onboarded_tenants(self, tenant_details):
        for idx in range(1, self.tenant_count + 1):
            details = tenant_details[idx - 1] if idx - 1 < len(tenant_details) else {}
            self.tenant_slots[idx]["owner_user"] = {
                "email": details.get("email"),
                "password": details.get("password"),
            }
        self._refresh_tenants_view()

    def _refresh_tenants_view(self):
        tenants_view = {}
        owners_view = {}
        for idx in range(1, self.tenant_count + 1):
            slot_data = self.tenant_slots.get(idx, {})
            tenant_data = slot_data.get("tenant", {})
            tenant_id = tenant_data.get("id") or idx
            owners = slot_data.get("owners", {})
            users = slot_data.get("users", {})

            owners_by_actual_id = {}
            for fallback_key, user_payload in owners.items():
                user_id = (user_payload or {}).get("id") or fallback_key
                owners_by_actual_id[user_id] = user_payload
                owners_view[user_id] = user_payload

            users_by_actual_id = {}
            for fallback_key, user_payload in users.items():
                user_id = (user_payload or {}).get("id") or fallback_key
                users_by_actual_id[user_id] = user_payload

            tenants_view[tenant_id] = {
                "tenant": {
                    "id": tenant_data.get("id"),
                    "name": tenant_data.get("name"),
                    "slug": tenant_data.get("slug"),
                },
                "owners": owners_by_actual_id,
                "users": users_by_actual_id,
                "roles": slot_data.get("roles", []),
            }
        self.tenants = tenants_view
        self.owners = owners_view

    def _store_tenant_users(self, tenant_index, tenant_user_details):
        users_map = {}
        fallback_idx = 1

        for item in (tenant_user_details or []):
            if isinstance(item, dict):
                user_id = item.get("id")
                users_map[user_id if user_id is not None else fallback_idx] = item
                fallback_idx += 1
                continue

            if isinstance(item, list):
                for user_item in item:
                    if not isinstance(user_item, dict):
                        continue
                    user_id = user_item.get("id")
                    users_map[user_id if user_id is not None else fallback_idx] = user_item
                    fallback_idx += 1

        self.tenant_slots[tenant_index]["users"] = users_map

    def _store_verified_users(self, tenant_login_details):
        for login_data in (tenant_login_details or []):
            if not isinstance(login_data, dict):
                continue

            user_data = login_data.get("user") or {}
            email = user_data.get("email")
            matched_idx = None

            for idx in range(1, self.tenant_count + 1):
                if self.tenant__owner_user_email(idx) == email:
                    matched_idx = idx
                    break

            if not matched_idx:
                continue

            tenant_info = (user_data.get("tenants") or [{}])[0]
            tenant_entry = self.tenant_slots[matched_idx]["tenant"]
            tenant_entry["id"] = tenant_info.get("id")
            tenant_entry["name"] = tenant_info.get("name")
            tenant_entry["slug"] = tenant_info.get("slug")

            owners_map = self.tenant_slots[matched_idx]["owners"]
            user_id = user_data.get("id")
            user_key = user_id if user_id is not None else (max(owners_map.keys()) + 1 if owners_map else matched_idx)
            owners_map[user_key] = {
                "id": user_data.get("id"),
                "email": user_data.get("email"),
                "password": self._owner_user_password(matched_idx),
                "access": login_data.get("access"),
                "refresh": login_data.get("refresh"),
                "user": user_data,
            }

        self._refresh_tenants_view()

    def tenant_onboarding(self):
        tenant_onboarding = TenantOnboarding(self.superuser_token)
        tenant_onboarding_test_cases, tenant_details = tenant_onboarding.start_test()
        self._store_onboarded_tenants(tenant_details)
        self.test_cases_sheets["Tenant Onboarding"] = tenant_onboarding_test_cases

    def tenant_email_verification(self):
        tenant_email_verification = TenantEmailVerification(
            email=self.tenant__owner_user_email(1),
            email2=self.tenant__owner_user_email(2)
        )

        tenant_email_verification_test_cases = tenant_email_verification.start_test()
        self.test_cases_sheets["Tenant Email Verification"] = tenant_email_verification_test_cases

    def tenant_request_otp(self):
        tenant_request_otp = RequestOTP(
            email=self.tenant__owner_user_email(1),
            email2=self.tenant__owner_user_email(2),
            password=self._owner_user_password(1)
        )

        tenant_request_otp_test_cases = tenant_request_otp.start_test()
        self.test_cases_sheets["Tenant Request OTP"] = tenant_request_otp_test_cases

    def tenant_validate_login_otp(self):
        tenant_validate_login_otp = ValidateOTP(
            email=self.tenant__owner_user_email(1),
            email2=self.tenant__owner_user_email(2),
            password=self._owner_user_password(1),
            otp=self.otp,
            otp2=self.otp
        )

        tenant_validate_login_otp_test_cases, tenant_login_details = tenant_validate_login_otp.start_test()
        self._store_verified_users(tenant_login_details)
        self.test_cases_sheets["Tenant Validate Login OTP"] = tenant_validate_login_otp_test_cases

    def tenant_roles(self):
        tenant_roles = TenantRoles(
            tenant_access_token=self._logged_in_tenant_owner_user_access_token(1),
            tenant2_access_token=self._logged_in_tenant_owner_user_access_token(2),
            tenant_id=self._tenant_id(1),
            tenant2_id=self._tenant_id(2)
        )

        tenant_roles_test_cases, tenant1_roles, tenant2_roles = tenant_roles.start_test()

        self.test_cases_sheets["Tenant Roles"] = tenant_roles_test_cases
        self.tenant_slots[1]["roles"] = tenant1_roles
        self.tenant_slots[2]["roles"] = tenant2_roles
        self._refresh_tenants_view()

    def tenant_users(self):
        tenant_users = TenantUsers(
            tenant_access_token=self._logged_in_tenant_owner_user_access_token(1),
            tenant2_access_token=self._logged_in_tenant_owner_user_access_token(2),
            tenant_id=self._tenant_id(1),
            tenant2_id=self._tenant_id(2),
            tenant_roles=self.tenant_slots[1].get("roles", []),
            tenant2_roles=self.tenant_slots[2].get("roles", [])
        )

        tenant_users_test_cases, tenant_user_details, tenant2_user_details = tenant_users.start_test()

        self.test_cases_sheets["Tenant Users"] = tenant_users_test_cases
        self._store_tenant_users(1, tenant_user_details)
        self._store_tenant_users(2, tenant2_user_details)
        self._refresh_tenants_view()



    def save_tests(self):
        filename = "test_cases.xlsx"
        self._export_excel(filename, self.test_cases_sheets)


    def start_test(self):
        self.tenant_onboarding()
        self.tenant_email_verification()
        self.tenant_request_otp()
        self.tenant_validate_login_otp()
        self.tenant_roles()
        self.tenant_users()
        self.save_tests()


def main():
    TestAutomation()


if __name__ == "__main__":
    main()
